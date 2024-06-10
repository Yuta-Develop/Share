from openpyxl import Workbook
import pandas as pd
import time
import datetime

date = datetime.date.today()

# ワークブックの新規作成と保存
wb = Workbook()
wb.save('Result.xlsx')

# ワークブックの読み込み
from openpyxl import load_workbook
wb = load_workbook('Result.xlsx')

# ワークシートの選択
ws = wb['Sheet']  # ワークシートを指定
ws = wb.active  # アクティブなワークシートを選択

ws.cell(row=2, column=2, value='Product')
ws.cell(row=2, column=3, value='ABC')
ws.cell(row=3, column=2, value='Phase')
ws.cell(row=3, column=3, value='TIB')
ws.cell(row=4, column=2, value='PWBA No.')
ws.cell(row=4, column=3, value='No.01')
ws.cell(row=5, column=2, value='Date')
ws.cell(row=5, column=3, value=date)
ws.cell(row=6, column=2, value='Remark')

ws.cell(row=9, column=5, value='Step')
ws.cell(row=9, column=6, value='Vin [Vac]')
ws.cell(row=9, column=7, value='Iin [Iac]')
ws.cell(row=9, column=8, value='Pin [Wac]')
ws.cell(row=9, column=9, value='Iout [Idc]')
ws.cell(row=9, column=10, value='Vout [Vdc]')
ws.cell(row=9, column=11, value='Pout [Wdc]')
ws.cell(row=9, column=12, value='η [%]')


df1 = pd.read_excel('Data.xlsx',sheet_name ='Data')
#print(df1)
df2 = df1.values.tolist()
#print(df2)
n=len(df2)

for i in range(n):
    num = df2[i][0]
    curma = df2[i][1]
    cura = curma / 1000
    
    row1 = i+10
    
    ws.cell(row=row1, column=5, value=num)
    ws.cell(row=row1, column=6, value=cura)
    
    
    
wb.save('Result.xlsx')